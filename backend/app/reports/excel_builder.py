"""Excel report builder using openpyxl."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_excel(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """Generate an in-memory Excel workbook and return raw bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Sheet name limit

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])

    # Auto-fit columns (approximate)
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
